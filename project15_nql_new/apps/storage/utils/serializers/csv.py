# -*- coding: utf-8 -*-

# @File   : csv
# @Author : Harold Chen
# @Date   : 2018/11/28

import os
from openpyxl import load_workbook
import csv
from apps.storage.models.data import DataMeta
from apps.storage.models.template import Template
from .flat_excel import FlatExcelWriter
from io import TextIOWrapper


def excel_to_csv(fp_or_path, dest_fp_or_path):
    """
    只支持导出第一张工作表
    :param fp_or_path:
    :param dest_fp_or_path:
    :return:
    """
    workbook = load_workbook(fp_or_path, data_only=True)
    worksheet = workbook.worksheets[0]
    if isinstance(dest_fp_or_path, str):
        dest_fp = open(dest_fp_or_path, 'w', encoding='utf-8')
    else:
        dest_fp = TextIOWrapper(dest_fp_or_path, encoding='utf-8', write_through=True)

    dest_fp.write('\ufeff')  # 辣鸡Excel
    wr = csv.writer(dest_fp, quoting=csv.QUOTE_ALL)

    for row in worksheet.iter_rows():
        l_row = []
        for cell in row:
            l_row.append(cell.value)
        wr.writerow(l_row)
    if isinstance(dest_fp, TextIOWrapper):
        dest_fp.detach()
    else:
        dest_fp.close()


class CSVWriter:

    def __init__(self, template: Template):
        self._flat_excel_writer = FlatExcelWriter(template)
        self._template = template

    def write_data(self, data_meta: DataMeta):
        self._flat_excel_writer.write_data(data_meta)

    def save(self, output_dir):
        filename = self._flat_excel_writer.save(output_dir)
        p = os.path.join(output_dir, filename)
        excel_to_csv(p, p)
        return filename
