import copy

import openpyxl as xl

from mgedata.generic.dataset import DataSet
from mgedata.generic.trans import TemplateContentToData
from .template import TemplateContentToXlsxDict
from .data import DataToXlsxDict, XlsxDictToData


class XlsxDictToXlsxFile:

    def __init__(self, xlsx_dict):
        self._xlsx_dict = xlsx_dict

    def _draw(self, wb, sheet_name, sheet_content):
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        values = [sheet_content['header']] + sheet_content['rows']
        for row in values:
            row_id, value_content = row[0], row[1:]
            for column_id in range(len(value_content)):
                ws.cell(row=row_id, column=column_id + 1).value = value_content[column_id]

    def _structure_to_sheet(self, structure):
        lists = []

        def _preorder(link_name, current_structure, path):
            nonlocal lists
            sheet_name, links = current_structure

            lists.append((len(lists) + 2, path, sheet_name))
            if links:
                for sub_link_name, sub_structure in links.items():
                    _preorder(sub_link_name, sub_structure, '.'.join((path, sub_link_name)))

        _preorder('Data Content', structure, 'content')
        return dict(
            header=(1, 'Template Structure', 'Sheet Name'),
            rows=lists
        )

    def to(self, place_holder: str) -> None:
        wb = xl.Workbook()
        wb.active.title = 'Data Info'

        self._draw(wb, 'Template Structure', self._structure_to_sheet(self._xlsx_dict['structure']))

        for sheet_name in self._xlsx_dict['ord']:
            self._draw(wb, sheet_name, self._xlsx_dict['values'][sheet_name])

        wb.save(place_holder)


class XlsxFileToXlsxDict:

    def __init__(self, place_holder: str):
        self._place_holder = place_holder

    def to(self) -> dict:
        ret = dict()

        wb = xl.load_workbook(self._place_holder)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            values = []
            for row_id, row in enumerate(ws.values, 1):
                value = tuple(cell for cell in row)
                if not all(x is None for x in value):
                    values.append((row_id,) + value)

            ret[sheet_name] = {
                'header': values[0],
                'rows': values[1:]
            }
        return dict(values=ret,
                    ord=wb.sheetnames)


class DataSetToXlsxDict:

    def __init__(self, data_set: DataSet):
        self._data_set = data_set

    def to(self) -> dict:
        xlsx_dict = TemplateContentToXlsxDict(self._data_set.template['content']).to()
        for data_id in self._data_set:
            xlsx_dict = DataToXlsxDict(self._data_set[data_id]).to(xlsx_dict)
        return xlsx_dict


class XlsxDictToDataSet:

    def __init__(self, xlsx_dict, template):
        self._xlsx_dict = xlsx_dict
        self._template = template

    @staticmethod
    def _sort_xlsx_dict(xlsx_dict):
        for sheet_name in xlsx_dict['ord'][1:]:
            xlsx_dict['values'][sheet_name]['rows'] = sorted(
                xlsx_dict['values'][sheet_name]['rows'],
                key=lambda x: x[1:]
            )
        return xlsx_dict

    def to(self):
        dataset = DataSet(template=self._template)
        xlsx_dict = copy.deepcopy(self._xlsx_dict)
        xlsx_dict = self._sort_xlsx_dict(xlsx_dict)

        while True:
            data = TemplateContentToData(self._template['content']).to()
            data = XlsxDictToData(data, xlsx_dict).to()
            if data is None:
                break
            dataset.add_data(data)
        return dataset
