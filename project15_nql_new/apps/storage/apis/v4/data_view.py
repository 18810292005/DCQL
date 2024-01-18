"""
对接 南大通用-数据观系统
"""
import logging
import os
import shutil
from enum import Enum, unique

import requests
from django.conf import settings
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from patoolib import extract_archive
from patoolib.util import PatoolError
from apps.sso_server.models import UserThirdPlatformToken
from apps.storage.models import DataMeta
from mgedata.errors.models import MGEError
from mgedata.utils.general import mkdtemp, require_POST_api, get_json_field_r, json_response
from apps.storage.models.file import FileContentType
from apps.storage.utils.data_import_export import export_data

logger = logging.getLogger('django')


@unique
class DataViewEnum(Enum):
    DATA_VIEW_TOKEN_URL = 'http://10.0.3.208:80/attach/openapi/token/usertoken'
    DATA_VIEW_UPLOAD_WORKBOOK_URL = "http://10.0.3.208:80/openapi/dataconnectors/store"
    DATA_VIEW_CREATE_SHEET_URL = "http://10.0.3.208:80/openapi/dataconnectors/excel/createtable?async=true"
    DATA_VIEW_CHECK_TOKEN = 'MftddE7DKpU3QTkXxIUxd8tIKXGnODVm'
    DATA_VIEW_ROOT_URL = 'http://dcs.nmdms.ustb.edu.cn/login'
    DATA_VIEW_REDIRECT_URL = "http://dcs.nmdms.ustb.edu.cn/api/oauth/gbase"


class DataView:
    def __init__(self, user):
        self.user = user
        self._token = ""
        self._data_field_id = ""
        self.get_user_token()

    @staticmethod
    def get_response_data(result: requests.Response, status_key, status_code):
        result.raise_for_status()
        result = result.json()
        msg = ''
        if result[status_key] != status_code:
            if result['msg'] is not None:
                msg = result['msg']
            raise MGEError.THIRD_PARTY_REQUEST_FAILED(msg)
        return result['data']

    def get_user_token(self):
        headers = {'token': DataViewEnum.DATA_VIEW_CHECK_TOKEN.value}
        result = requests.get(DataViewEnum.DATA_VIEW_TOKEN_URL.value, headers=headers,
                              params={'email': self.user.email})
        response_data = self.get_response_data(result, 'code', 200)
        if response_data['token'] is None:
            raise MGEError.THIRD_PARTY_REQUEST_FAILED('用户未打开数据观-ETL数据导入功能')
        self._token = response_data['token']

    def upload_workbook_to_data_view(self, file_path):
        headers = {"Authorization": f"OAuth {self._token}"}
        with open(file_path, 'rb') as f:
            result = requests.post(DataViewEnum.DATA_VIEW_UPLOAD_WORKBOOK_URL.value, headers=headers,
                                   files={'files[]': f})
            response_data = self.get_response_data(result, 'status', 0)
            self._data_field_id = response_data['storageDataId']
        return response_data

    def create_sheets(self, file_path):
        file_name = os.path.basename(file_path)
        basename, ext = os.path.splitext(file_name)
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            sheet_index = wb.index(sheet)
            excel_index_name = f"{basename}-{sheet_name}"
            headers = {"Authorization": f"OAuth {self._token}", "Content-Type": "application/json;charset=UTF-8"}
            data = {"sourceType": "EXCEL", "dataFileId": self._data_field_id, "sheetIndex": sheet_index,
                    "titleRow": 0,
                    "name": excel_index_name}
            result = requests.post(DataViewEnum.DATA_VIEW_CREATE_SHEET_URL.value, headers=headers, json=data)
            self.get_response_data(result, 'status', 0)
            # TODO:记录用户，上传文件信息，是否成功，上传时间

    def get_upload_excel_file_list(self, archive_path):
        # 获取压缩包文件的绝对路径
        archive_path = os.path.join(settings.MEDIA_ROOT, archive_path).replace(settings.MEDIA_URL, '/')
        tmp_dir = mkdtemp()
        tmp_dir_unarchived = os.path.join(tmp_dir, 'unarchived')
        os.mkdir(tmp_dir_unarchived)
        try:
            list_path = []
            extract_archive(archive_path, outdir=tmp_dir_unarchived)
            for root, dirs, files in os.walk(tmp_dir_unarchived):
                for f in files:
                    basename, ext = os.path.splitext(f)
                    ext = ext.upper()
                    if ext == '.XLSX' and not basename.startswith('~'):
                        list_path.append(os.path.join(root, f))
            return list_path
        except PatoolError as ex:
            logging.error(str(ex))
            shutil.rmtree(tmp_dir)
            raise MGEError.BAD_ARCHIVE


@require_POST_api
@login_required
def upload_export_excel_to_data_view(request):
    data_meta_list = get_json_field_r(request, 'meta_ids', list)
    meta_list = DataMeta.objects.filter(id__in=data_meta_list)

    user = request.user
    if meta_list.count() == 0:
        raise MGEError.NO_AVAILABLE_DATA
    data_view = DataView(user)
    archive_media_path: str = export_data(user=user, meta_or_list=meta_list, file_format=FileContentType.XLSX,
                                          asynchronous=False, no_attachments=True)[1:]
    excel_paths: list = data_view.get_upload_excel_file_list(archive_media_path)
    for excel_path in excel_paths:
        data_view.upload_workbook_to_data_view(excel_path)
        data_view.create_sheets(excel_path)

    try:
        token = UserThirdPlatformToken.objects.get(user=user, third_platform__platform_name='ndty-ssologin').token
        redirect_url = DataViewEnum.DATA_VIEW_REDIRECT_URL.value + f'?token={token}'
    except Exception:
        redirect_url = DataViewEnum.DATA_VIEW_ROOT_URL.value
    return json_response({'redirect_url': redirect_url})
