# -*- coding: utf-8 -*-

# @File   : tasks.py
# @Author : Harold Chen
# @Date   : 2018/1/17

from __future__ import absolute_import, unicode_literals

import json
import logging
import os
import random
import shutil
import traceback
from datetime import datetime
from hashlib import md5

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import requests
from bson.objectid import ObjectId
from celery import shared_task
from django.conf import settings
from django.db import models
from django.db import transaction
from django.db.models import Q, Count
from django.utils import timezone
from django.utils.translation import ugettext as _
from patoolib import extract_archive
from patoolib.util import PatoolError

from apps.account.models import User
from apps.account.notify import NotificationType, notify
from apps.certificate.models import TemplateStatistic
from apps.mge.models import PersistentVariables, PVEnum
from apps.search.tasks import import_to_es
from apps.storage.models.data import DataMeta, DataUploadSourceMap, DataSet, DataReviewState
from apps.storage.models.data import UploadHistory
from apps.storage.models.file import FileContentType, TemporaryExportedFile
from apps.storage.models.file import UploadedSourceFile, TemporaryUploadedFile
from apps.storage.models.material import MaterialCategory
from apps.storage.models.material import MaterialProject
from apps.storage.models.template import Template
from apps.storage.utils.constant import CONST
from apps.storage.utils.serializers import (
    DataMode, JSONHandler, ExcelHandler, XMLHandler, FlatJSONWriter, FlatExcelWriter, OCPMDMSerializer, CSVWriter
)
from apps.storage.utils.serializers.common import ParsingError
from apps.task.task_registry import MGEBaseTask
from mgedata.errors.models import MGEError
from mgedata.test.utils import instances
from mgedata.utils.general import mkdtemp, mge_make_archive
from mgedata.task_log import mge_task_log

logger = logging.getLogger('django')


class DataImportTask(MGEBaseTask):

    @property
    def task_description(self):
        return "数据汇交"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._meta_id_list = []
        self._to_be_revoked = []
        self._history_id = None

    def run(self, temporary_uploaded_file_id, file_format, uploaded_by, verify_only=False):
        super().run()
        file_format = FileContentType(file_format)  # file_format被celery传递时会被专为int，要转回来
        mge_task_log("开始数据汇交", {"temporary_uploaded_file_id": temporary_uploaded_file_id,
        "file_format": file_format, "uploaded_by": uploaded_by, "verify_only": verify_only})
        try:
            temp_file = TemporaryUploadedFile.objects.get(id=temporary_uploaded_file_id)
            temp_file.file.seek(0)
            fp = temp_file.file
            fp.name = temp_file.filename
        except (TemporaryUploadedFile.DoesNotExist, FileNotFoundError):
            raise MGEError.TASK_FILE_EXPIRED

        def _unarchive_files(archive_fp):
            mge_task_log("开始解压", {"filename": archive_fp.name})
            tmp_dir = mkdtemp()
            tmp_archive_file_path = os.path.join(tmp_dir, 'original_archive')
            tmp_unarchived_dir_path = os.path.join(tmp_dir, 'unarchived')
            os.mkdir(tmp_unarchived_dir_path)
            mge_task_log("开始移动压缩文件到", {"filepath": tmp_archive_file_path})
            with open(tmp_archive_file_path, 'wb') as f:  # 暂存用户上传的文件
                shutil.copyfileobj(archive_fp, f)
            try:
                extract_archive(tmp_archive_file_path, outdir=tmp_unarchived_dir_path)
                parser_file = None
                for root, dirs, files in os.walk(tmp_unarchived_dir_path):
                    if '__MACOSX' in root:
                        continue
                    for f in files:
                        basename, ext = os.path.splitext(f)
                        try:  # 如果源压缩包名字是中文，名字是cp437格式编码的会存在乱码,需要重命名文件
                            new_filename = basename.encode('cp437').decode('gbk') + ext
                        except Exception:
                            new_filename = basename + ext  # 如果不是cp437编码,不做处理
                        new_filename = new_filename.strip()  # 去除文件名开头结尾空格
                        source_file_path = os.path.join(root, f)
                        target_file_path = os.path.join(root, new_filename)
                        os.rename(source_file_path, target_file_path)
                        ext = ext.upper()
                        if ext in ('.XLSX', '.JSON', '.XML') and parser_file is None:
                            parser_file = root
                if parser_file is None:
                    raise MGEError.NO_VALID_FILE
                return parser_file

            except PatoolError as ex:
                shutil.rmtree(tmp_dir)
                mge_task_log("PatoolError", {"error": ex}, level = logging.ERROR)
                logger.error(ex)
                raise MGEError.BAD_ARCHIVE

        def _filter_files(output_path: str) -> dict:
            """
            用户上传的文件分类，分为XLSX，JSON，XML文件
            :param output_path: 文件的所在文件夹路径
            :return:
            """
            mge_task_log("用户上传文件分类", {"output_path": output_path})
            file_names = os.listdir(output_path)
            xlsx_candidates = []
            json_candidates = []
            xml_candidates = []
            for sub_filename in file_names:
                basename, ext = os.path.splitext(sub_filename)
                ext = ext.upper()
                if ext == '.XLSX':
                    if basename.startswith('~'):  # 过滤excel临时文件
                        continue
                    xlsx_candidates.append(sub_filename)
                elif ext == '.JSON':
                    json_candidates.append(sub_filename)
                elif ext == '.XML':
                    xml_candidates.append(sub_filename)

            if len(xlsx_candidates) + len(json_candidates) + len(xml_candidates) == 0:
                raise MGEError.NO_VALID_FILE
            return {
                FileContentType.XLSX: xlsx_candidates,
                FileContentType.JSON: json_candidates,
                FileContentType.XML: xml_candidates
            }

        if file_format == FileContentType.ARCHIVE:
            try:
                temp_dir = _unarchive_files(fp)
            except PatoolError as e:
                logger.error(e)
                raise MGEError.BAD_ARCHIVE
        else:
            temp_dir = mkdtemp()
            temp_file_path = os.path.join(temp_dir, fp.name)
            with open(temp_file_path, 'wb') as temp_fp:
                shutil.copyfileobj(fp, temp_fp)

        try:
            candidates = _filter_files(temp_dir)
            count = 0
            total_count = len(candidates[FileContentType.XLSX]) + len(candidates[FileContentType.XML]) + len(
                candidates[FileContentType.JSON])
            if len(candidates) == 0:
                raise MGEError.NO_AVAILABLE_DATA

            def progress_handler(percentage, block=False):
                self.update_progress((percentage + count) / total_count, block)
            mge_task_log("开始数据读取")
            for f_format, filename_list in candidates.items():
                for filename in filename_list:
                    if f_format.is_xlsx:
                        handler = ExcelHandler(DataMode.READ, verify_only=verify_only)
                    elif f_format.is_json:
                        handler = JSONHandler(DataMode.READ, verify_only=verify_only)
                    else:
                        handler = XMLHandler(DataMode.READ, verify_only=verify_only)
                    self._meta_id_list.extend(
                        handler.read_data(os.path.join(temp_dir, filename), file_dir=temp_dir, uploaded_by=uploaded_by,
                                          progress_handler=progress_handler))
                    count += 1
            if not verify_only:
                with transaction.atomic():
                    file = UploadedSourceFile.add(fp, content_type=file_format, filename=fp.name)
                    [DataUploadSourceMap(meta_id=meta_id, file=file).save() for meta_id in self._meta_id_list]
                    DataMeta.importing_objects.filter(id__in=self._meta_id_list).update(importing=False)

                    # 数据集
                    if len(self._meta_id_list) != 0:
                        user = User.objects.filter(username=uploaded_by)[0]
                        data_meta = DataMeta.objects.get(id=self._meta_id_list[0])
                        title = data_meta.title
                        project = data_meta.get_project_id()
                        contributor = data_meta.contributor
                        data_set = DataSet.objects.create(title=title, project=project, contributor=contributor,
                                                          user=user)
                        DataMeta.objects.filter(pk__in=self._meta_id_list).update(dataset=data_set)

                    h = UploadHistory(user_id=uploaded_by, source=file, meta_id_list=self._meta_id_list, via_file=True,
                                      count=len(self._meta_id_list),
                                      category=handler.template.category)
                    h.save()
                    self._history_id = h.pk
                mge_task_log("开始同步到es", {"meta_id_list": self._meta_id_list, "count": len(self._meta_id_list)})
                import_to_es.delay(self._meta_id_list)
                return {'data_count': len(self._meta_id_list), 'meta_id_list': self._meta_id_list}
        except ParsingError as e:
            if not verify_only:
                DataMeta.importing_objects.filter(id__in=self._meta_id_list).delete()
            raise MGEError.BAD_DATA(e.message)
        except ValueError as e:
            if not verify_only:
                DataMeta.importing_objects.filter(id__in=self._meta_id_list).delete()
            raise MGEError.BAD_DATA(str(e))
        finally:
            self._to_be_revoked = self._meta_id_list[:]
            self._meta_id_list.clear()
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                logger.error(e)

    def on_success_but_revoked(self):
        if self._history_id:
            with transaction.atomic():
                DataMeta.objects.filter(id__in=self._to_be_revoked).delete()
                UploadHistory.objects.filter(pk=self._history_id).delete()


class DataExportTask(MGEBaseTask):
    @property
    def task_description(self):
        return "数据导出"

    def run(self, username, meta_id_list, file_format: FileContentType, excluded_fields=(), included_fields=(),
            no_attachments=False, flat=False, asynchronous=True):
        try:
            user = User.objects.get(username = username)
        except User.DoesNotExist:
            return MGEError.USER_NOT_FOUND("username不存在")

        super().run()
        file_format = FileContentType(file_format)  # file_format被celery传递时会被专为int，要转回来
        if len(meta_id_list) == 0:
            raise MGEError.NO_AVAILABLE_DATA
        
        public_id_list = DataMeta.objects.filter(pk__in=meta_id_list).filter(public_range = "public").values_list("id", flat= True)
        project_id_list = DataMeta.objects.filter(pk__in = meta_id_list).filter(public_range = "project").filter(other_info__project__in = user.projects_list).values_list("id", flat= True)
        subject_id_list = DataMeta.objects.filter(pk__in = meta_id_list).filter(public_range = "subject").filter(other_info__subject__in = user.subjects_list).values_list("id", flat=True)
        private_id_list = DataMeta.objects.filter(pk__in = meta_id_list).filter(public_range = "private").filter(user = user).values_list("id", flat=True)

        meta_id_list = list(set(public_id_list) | set(project_id_list) | set(subject_id_list) | set(private_id_list))
        tid_handler_map = {}
        tid_list = [x['template_id'] for x in
                    DataMeta.objects.filter(pk__in=meta_id_list).values('template_id').order_by(
                        'template_id').distinct()]
        if not tid_list:
            raise MGEError.NO_AVAILABLE_DATA
        elif len(tid_list) > 1 and (excluded_fields or included_fields):
            raise MGEError.UNSUPPORTED_FIELDS_FILTERING(_("Multiple templates involved"))
        for tid in tid_list:
            template = Template.objects.get(id=tid)
            if excluded_fields or included_fields:
                template.filter_fields(excluded_fields=excluded_fields, included_fields=included_fields)
            if file_format == FileContentType.XLSX:
                if flat:
                    handler = FlatExcelWriter(template=template)
                else:
                    handler = ExcelHandler(DataMode.WRITE, template=template, no_attachments=no_attachments)
            elif file_format == FileContentType.JSON:
                if flat:
                    handler = FlatJSONWriter(template=template, with_headers=True)
                else:
                    handler = JSONHandler(DataMode.WRITE, template=template, no_attachments=no_attachments)
            elif file_format == FileContentType.XML:
                handler = XMLHandler(DataMode.WRITE, template=template, no_attachments=no_attachments)
            elif file_format == FileContentType.CSV:
                handler = CSVWriter(template=template)
            else:
                handler = None  # NEVER
            tid_handler_map[tid] = handler

        if isinstance(meta_id_list, models.QuerySet):
            meta_list = meta_id_list
            total_count = meta_list.count()
        elif len(meta_id_list) > 0 and isinstance(meta_id_list[0], DataMeta):
            meta_list = meta_id_list
            total_count = len(meta_list)
        else:
            meta_list = DataMeta.objects.filter(id__in=meta_id_list)
            total_count = len(meta_list)

        current_count = 0

        for meta in meta_list:
            handler = tid_handler_map[meta.template_id]
            handler.write_data(meta)
            current_count += 1
            if current_count % 3 == 0:
                # 每3条数据更新一次进度
                self.update_progress(current_count / total_count)
        # if len(tid_handler_map) == 1:
        #     handler = list(tid_handler_map.values())[0]
        #     fp = BytesIO()
        #     filename = handler.save(fp)
        #     return {
        #         'url': TemporaryExportedFile.add(fp, content_type=file_format, filename=filename,
        #                                          randomize=False).get_url()}

        temp_dir = None
        try:
            temp_dir = mkdtemp()
            file_dir = os.path.join(temp_dir, 'files')
            os.makedirs(file_dir, exist_ok=True)
            archive_name = f'{timezone.now().strftime("%Y%m%d%H%M%S")}'
            archive = os.path.join(temp_dir, archive_name)
            self.update_progress(-1, block=True)
            for tid, handler in tid_handler_map.items():
                handler.save(file_dir).replace('/', '／').replace('\\', '＼')
            mge_make_archive(archive, 'zip', root_dir=file_dir)
            archive += '.zip'
            archive_name += '.zip'
            with open(archive, 'rb') as f:
                return {'url': TemporaryExportedFile.add(f, content_type=FileContentType.ARCHIVE, filename=archive_name,
                                                         randomize=False).get_url()}
        finally:
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                logger.error(e)


class DataExportForCertificate(MGEBaseTask):

    def run(self, ps_id: str, is_project: bool, template_id: int, random_percentage: float = 1):
        assert 1 >= random_percentage > 0
        super().run()
        try:
            research_cat = MaterialCategory.objects.get(name_zh='科研成果')
        except MaterialCategory.DoesNotExist:
            raise MGEError.NOT_FOUND("找不到科研成果模板，请联系管理员")
        try:
            template = Template.objects.get(id=template_id)
        except Template.DoesNotExist:
            raise MGEError.NOT_FOUND(f"模板{template_id}不存在")
        is_research_template = template.category_id == research_cat.id
        no_attachments = True
        if not is_research_template:
            # 如果非科研成果，判断模板是否属于项目
            if TemplateStatistic.objects.filter(template_id=template_id).count() == 0:
                name = "项目" if is_project else "课题"
                raise MGEError.PERMISSION_DENIED(f"模板{template_id}不属于{name}{ps_id}")
            if random_percentage == 1:
                no_attachments = True
            else:
                no_attachments = False
        else:
            no_attachments = False
        handler = ExcelHandler(DataMode.WRITE, template=template,
                               no_attachments=no_attachments)

        if is_project:
            _other_info_q = Q(other_info__project=ps_id)
        else:
            _other_info_q = Q(other_info__subject=ps_id)
        q = Q(template_id=template_id, review_state=DataReviewState.APPROVED.value)
        meta_list = DataMeta.objects.filter(q & _other_info_q)
        total_count = meta_list.count()
        if total_count == 0:
            raise MGEError.NO_AVAILABLE_DATA("请重新生成汇交证明")  # certificate统计有误

        if random_percentage < 1:
            total_count *= random_percentage
            total_count = max(int(total_count), 1)
            meta_list = list(meta_list.values_list('id', flat=True))
            rand = random.sample(meta_list, total_count)
            if len(rand) > 1000:
                rand = rand[:1000]
            meta_list = DataMeta.objects.filter(id__in=rand)
        current_count = 0
        for meta in meta_list:
            handler.write_data(meta)
            current_count += 1
            if current_count % 3 == 0:
                # 每3条数据更新一次进度
                self.update_progress(current_count / total_count)
        temp_dir = None
        self.update_progress(-1, block=True)
        try:
            temp_dir = mkdtemp()
            file_dir = os.path.join(temp_dir, 'files')
            os.makedirs(file_dir, exist_ok=True)
            archive_name = f'{timezone.now().strftime("%Y%m%d%H%M%S")}'
            archive = os.path.join(temp_dir, archive_name)
            handler.save(file_dir)
            mge_make_archive(archive, 'zip', root_dir=file_dir)
            archive += '.zip'
            archive_name += '.zip'
            with open(archive, 'rb') as f:
                return {'url': TemporaryExportedFile.add(f, content_type=FileContentType.ARCHIVE, filename=archive_name,
                                                         randomize=False).get_url()}
        finally:
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                logger.error(e)


class OCPMDMDataExportTask(MGEBaseTask):
    _NOTICE_URL = 'https://ocpmdm.mgedata.cn/export_notice'
    _SEND_URL = 'https://ocpmdm.mgedata.cn/store_export'

    def run(self, username, meta_id_list, filename=None, excluded_fields=(), included_fields=()):
        super().run()
        # try:
        #     task = UserTask.objects.get(celery_id=str(self.request.id))
        #     exported_file_id = (task.result or {}).get('file_id')
        #     noticed_start = (task.result or {}).get('noticed_start')
        #     noticed_end = (task.result or {}).get('noticed_end')
        #     sent = (task.result or {}).get('sent')
        #     if exported_file_id:
        #         try:
        #             file = TemporaryExportedFile.objects.get(id=exported_file_id)
        #             result = json.loads(file.file, encoding='utf-8')
        #             export_id = result['export_id']
        #             params = result['params']
        #             json_string = result['data']
        #             print(noticed_start, noticed_end)
        #             if not noticed_start:
        #                 self.notice_ocpmdm(export_id, username, filename)
        #             if not noticed_end:
        #                 self.notice_ocpmdm(export_id, username, filename, params)
        #             if not sent:
        #                 self.send_to_ocpmdm(export_id, json_string)
        #         except TemporaryExportedFile.DoesNotExist:
        #             pass
        # except UserTask.DoesNotExist:
        #     pass
        if not filename:
            filename = f'{username}_{datetime.now().strftime("%Y%m%d%H%M%S")}'

        if len(meta_id_list) == 0:
            raise MGEError.NO_AVAILABLE_DATA
        tid_list = [x['template_id'] for x in
                    DataMeta.objects.filter(pk__in=meta_id_list).values('template_id').order_by(
                        'template_id').distinct()]
        if not tid_list:
            raise MGEError.NO_AVAILABLE_DATA
        elif len(tid_list) > 1:
            raise MGEError.UNSUPPORTED_FIELDS_FILTERING(_("Multiple templates involved"))
        tid = tid_list[0]
        template = Template.objects.get(id=tid)
        if excluded_fields or included_fields:
            template.filter_fields(excluded_fields, included_fields)

        object_id = ObjectId()
        m = md5()
        m.update(str(object_id).encode())
        export_id = m.hexdigest()[8:-8]

        try:
            # 导出之前通知
            self.notice_ocpmdm(export_id, username, filename)
        except Exception:
            raise MGEError.THIRD_PARTY_COMMUNICATION_FAILED(_("Failed to connect to OCPMDM."))

        handler = OCPMDMSerializer(template=template)

        if isinstance(meta_id_list, models.QuerySet):
            meta_list = meta_id_list
            total_count = meta_list.count()
        elif len(meta_id_list) > 0 and isinstance(meta_id_list[0], DataMeta):
            meta_list = meta_id_list
            total_count = len(meta_list)
        else:
            meta_list = DataMeta.objects.filter(id__in=meta_id_list)
            total_count = len(meta_list)

        current_count = 0
        for meta in meta_list:
            handler.write_data(meta)
            current_count += 1
            if current_count % 3 == 0:
                # 每3条数据更新一次进度
                self.update_progress(current_count / total_count)

        data, headers, header_type_map = handler.save()

        params = [[header, header_type_map[header]] for header in headers]
        result = {
            'export_id': export_id,
            'params': params,
            'data': json.dumps(data, ensure_ascii=False)
        }
        try:
            # 发送数据
            self.send_to_ocpmdm(export_id, result['data'])
            # 导出之后通知
            self.notice_ocpmdm(export_id, username, filename, params)
        except Exception:
            raise MGEError.THIRD_PARTY_COMMUNICATION_FAILED(_("Failed to connect to OCPMDM."))
        return filename

    def notice_ocpmdm(self, export_id, username, filename, params=None):
        post_json = {
            'export_id': export_id,
            'file_name': filename,
            'user_name': username,
            'type': 'start'
        }
        if params:
            post_json['param_info'] = json.dumps(params, ensure_ascii=False)
            post_json['type'] = 'end'
            print(post_json['param_info'])
        print("notice " + post_json['type'] + " export_id:" + export_id)
        self._post_request(OCPMDMDataExportTask._NOTICE_URL, post_json)

    def send_to_ocpmdm(self, export_id, json_string: str):
        post_json = {
            'export_id': export_id,
            'data': json_string
        }
        print("sending to ocpmdm")
        self._post_request(OCPMDMDataExportTask._SEND_URL, post_json)

    def _post_request(self, url, data_dict):
        request = requests.post(url=url, json=data_dict)
        print(request.content)


class ProjectInFoExportTask(MGEBaseTask):

    @property
    def task_description(self):
        return "统计信息Excel导出"

    def run(self, projects_dict, project_header_name):
        super().run()
        # 可以传入id或者传入map
        if len(projects_dict) == 0 or len(project_header_name) == 0:
            raise MGEError.NO_AVAILABLE_DATA

        wb = Workbook()
        ws = wb.active
        # Excel生成日期.
        ws.append([f"excel生成日期:{timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.merge_cells(f"A1:{get_column_letter(len(project_header_name))}1")
        ws.append(project_header_name)
        for project in projects_dict:
            ws.append(tuple(project.values()))
        # 设置列宽
        for t in range(len(project_header_name)):
            if t == 1:
                ws.column_dimensions[get_column_letter(t + 1)].width = 70  # 项目编号70宽度
            else:
                ws.column_dimensions[get_column_letter(t + 1)].width = 15
        # 设置表格居中状态
        for i in range(ws.max_row):
            for j in range(ws.max_column):
                from openpyxl.styles import Alignment
                ws.cell(row=i + 1, column=j + 1).alignment = Alignment(horizontal='center', vertical='center')
        try:
            temp_dir = mkdtemp()
            file_dir = os.path.join(temp_dir, 'files')
            os.makedirs(file_dir, exist_ok=True)
            archive_name = f'项目统计信息:{timezone.now().strftime("%Y%m%d %H%M%S")}.xlsx'
            archive = os.path.join(file_dir, archive_name)
            wb.save(archive)
            self.update_progress(-1, block=True)
            with open(archive, 'rb') as f:
                return {'url': TemporaryExportedFile.add(f, content_type=FileContentType.XLSX, filename=archive_name,
                                                         randomize=False).get_system_path(absolute=True)}
        finally:
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                logger.error(e)


@shared_task
def material_project_syn():
    """
    同步项目和课题
    """
    if PersistentVariables.get_variable(PVEnum.MATERIAL_PROJECT_LOCK.value):
        PersistentVariables.set_variable(PVEnum.MATERIAL_PROJECT_LOCK.value, False)
        try:
            if settings.DEBUG:  # 测试时更新
                remote_version = None
            else:
                remote_version = requests.get(url=settings.MATERIAL_PROJECT_VERSION_URL).json()
            if settings.DEBUG:
                material_projects = instances.material_projects
            else:
                material_projects = requests.get(url=settings.MATERIAL_PROJECT_GET_URL).json()["data"]
            local_version = PersistentVariables.get_variable(PVEnum.MATERIAL_PROJECT_VERSION.value)
            # 同步
            if remote_version is None or remote_version > local_version:
                MaterialProject.make_for_syn(material_projects, remote_version)
        except Exception:
            print(f'error material_project_syn:{traceback.format_exc()}')
        PersistentVariables.set_variable(PVEnum.MATERIAL_PROJECT_LOCK.value, True)


@shared_task
def register_doi_task(data_ids, doi):
    """
    大量数据的doi赋值任务
    """
    for data_id in data_ids:
        try:
            data = DataMeta.objects.get(id=data_id)
            if data.doi is None or data.doi == '':
                data.doi = doi
                data.save()
        finally:
            print(f"register_doi_task error, info: {traceback.format_exc()}")
    import_to_es(data_ids)


class UploadHistoryReviewRevokeTask(MGEBaseTask):
    @property
    def task_description(self):
        return "数据集审核撤回"

    def run(self, history_id):
        def cut(obj, sec):
            return [obj[i:i + sec] for i in range(0, len(obj), sec)]

        super().run()
        try:
            history = UploadHistory.objects.get(id=history_id)
            if history.review_state == DataReviewState.PENDING:
                raise MGEError.BAD_DATA("数据未审核，拒绝撤回！")
            max_slice = 1000
            current_count = 0
            total_count = int(len(history.meta_id_list) / max_slice) + 1
            with transaction.atomic():
                meta_id_slice_list = cut([i for i in history.meta_id_list], max_slice)
                for meta_id_slice in meta_id_slice_list:
                    DataMeta.objects.filter(id__in=meta_id_slice).update(review_state=DataReviewState.PENDING,
                                                                         reviewer=None)
                    current_count += 1
                    self.update_progress(current_count / total_count)
                UploadHistory.objects.filter(id=history_id).update(review_state=DataReviewState.PENDING, reviewer=None,
                                                                   disapprove_reason=None)
            notify(history.user, NotificationType.DATA_REVIEW_REVOKE, history.format_time)
            return {'upload_user': history.user.username, 'time': history.format_time}
        except Exception as e:
            raise e


class UploadHistoryRetractTask(MGEBaseTask):
    @property
    def task_description(self):
        return "数据集撤回"

    def run(self, history_id):
        super().run()
        try:
            history = UploadHistory.objects.get(id=history_id)
            with transaction.atomic():
                total_count = len(history.meta_id_list)
                DataMeta.objects.filter(id__in=history.meta_id_list).delete()
                history.delete()
                import_to_es(history.meta_id_list)
                notify(history.user, NotificationType.DATA_RETRACT, history.format_time, history_id)
            return {'data_count': total_count}
        except Exception as e:
            raise e


class RebuildFileUsageTask(MGEBaseTask):
    """
    重建数据与图片、文件的引用关系
    """

    def run(self, meta_id_list):
        super().run()
        from apps.storage.docs import DataContent, DataFieldContainer, DataFieldTableRow
        from apps.storage.models.file import FileUsage, ImageUsage, DataContentFile, DataContentImage
        from django.conf import settings
        if not meta_id_list:
            return
        res = {}  # 统计结果，key是文件路径，value是使用此文件的文件数组

        for table in DataContent, DataFieldContainer, DataFieldTableRow:
            map_func = """
                        function () {
                            res = []
                            for (var key in this) {
                                value = this[key];
                                if (Array.isArray(value)){
                                    for (var ele of value){
                                        if (typeof(ele)=='string' && ele.startsWith('_fs')){
                                            res.push(ele);
                                        }
                                    }
                                }
                                else if (typeof(value)=='string' && value.startsWith('_fs')){
                                    res.push(value);
                                }
                            }
                            if (res.length != 0){
                                emit(this._meta_id, res);
                            }
                        }
           """
            reduce_func = """
                function (key, values) {
                    var res = [];
                    for (value of values){
                       if(Array.isArray(value)){
                           for (v of value){
                               res.push(v);
                           }
                       }
                       else{
                           res.push(value);
                       }
                    }
                    return {'res':res};
                }
            """
            self.update_progress(0.01, extra=f'Map-reducing {table.__name__}')
            mr = table.objects(_meta_id__in=meta_id_list).map_reduce(map_f=map_func, reduce_f=reduce_func,
                                                                     output=f'rebuild_{table.__name__}')
            for doc in mr:
                if isinstance(doc.value, list):
                    files = doc.value
                else:
                    files = doc.value['res']
                for file in list(files):
                    if isinstance(file, dict):
                        for _file in file['res']:
                            res.setdefault(_file, []).append(int(doc.key))
                    else:
                        res.setdefault(file, []).append(int(doc.key))

        total_count = len(res)
        if total_count == 0:
            return

        self.update_progress(0.01, extra='统计缺失文件')
        file_paths = list(res.keys())
        q_image = list(DataContentImage.objects.filter(file__in=file_paths).values_list('file', 'id', 'size'))
        q_file = list(DataContentFile.objects.filter(file__in=file_paths).values_list('file', 'id', 'size'))
        not_found = list(set(file_paths) - (set([x[0] for x in q_image]) | set([x[0] for x in q_file])))
        for file, _id, size in q_image:
            try:
                if size == 0:
                    size = os.path.getsize(os.path.join(settings.MEDIA_ROOT, file))
                    DataContentImage.objects.filter(pk=_id).update(size=size)
            except FileNotFoundError:
                pass
        for file, _id, size in q_file:
            try:
                if size == 0:
                    size = os.path.getsize(os.path.join(settings.MEDIA_ROOT, file))
                    DataContentFile.objects.filter(pk=_id).update(size=size)
            except FileNotFoundError:
                pass
        _len = len(not_found)
        data_images = []
        data_image_paths = []
        data_files = []
        data_file_paths = []
        for i, file_path in enumerate(not_found):
            full_path = os.path.join(settings.MEDIA_ROOT, file_path)
            try:
                size = os.path.getsize(full_path)
            except FileNotFoundError:
                continue
            if os.path.splitext(file_path)[1].upper() in CONST.SUPPORT_IMAGE_EXT:
                data_images.append(
                    DataContentImage(file=file_path, size=size, hash_value='rebuild')
                )
                data_image_paths.append(file_path)
            else:
                data_files.append(
                    DataContentFile(file=file_path, size=size, hash_value='rebuild')
                )
                data_file_paths.append(file_path)
            if i % min(10, _len) == 0:
                self.update_progress(min((i + 1) / _len, 0.95), extra=f'添加文件 {i}/{_len}')
        self.update_progress(0.96, extra='更新DataContentFile')
        DataContentFile.objects.bulk_create(data_files)
        self.update_progress(0.97, extra='更新DataContentImage')
        DataContentImage.objects.bulk_create(data_images)
        q_image.extend(
            list(DataContentImage.objects.filter(file__in=data_image_paths).values_list('file', 'id', 'size')))
        q_file.extend(list(DataContentFile.objects.filter(file__in=data_file_paths).values_list('file', 'id', 'size')))
        q_image_len = len(q_image)
        q_file_len = len(q_file)
        file_bulk = []
        image_bulk = []

        image_usage_pairs = []
        file_usage_pairs = []

        for i, (file_path, file_id, _) in enumerate(q_image):
            meta_ids = res[file_path]
            dummy_file_id_list = [file_id] * len(meta_ids)
            image_usage_pairs.extend(list(zip(meta_ids, dummy_file_id_list)))
            self.update_progress(0.98, extra=f'统计图片引用{i + 1}/{q_image_len}')

        for i, (file_path, file_id, _) in enumerate(q_file):
            meta_ids = res[file_path]
            dummy_file_id_list = [file_id] * len(meta_ids)
            file_usage_pairs.extend(list(zip(meta_ids, dummy_file_id_list)))
            self.update_progress(0.98, extra=f'统计文件引用{i + 1}/{q_file_len}')

        self.update_progress(0.99, extra='更新图片引用')
        for i, (meta_id, file_id) in enumerate(image_usage_pairs):
            image_bulk.append(ImageUsage(meta_id=meta_id, file_id=file_id))
            if i % 2000 == 0:
                self.update_progress(0.99, extra=f'更新图片引用{i + 1}/{len(image_usage_pairs)}')
                ImageUsage.objects.bulk_create(image_bulk, ignore_conflicts=True)
                image_bulk = []
        self.update_progress(0.99, extra='更新文件引用')
        for i, (meta_id, file_id) in enumerate(file_usage_pairs):
            file_bulk.append(FileUsage(meta_id=meta_id, file_id=file_id))
            if i % 2000 == 0:
                self.update_progress(0.99, extra=f'更新文件引用{i + 1}/{len(file_usage_pairs)}')
                FileUsage.objects.bulk_create(file_bulk, ignore_conflicts=True)
                file_bulk = []
        ImageUsage.objects.bulk_create(image_bulk, ignore_conflicts=True)
        FileUsage.objects.bulk_create(file_bulk, ignore_conflicts=True)
