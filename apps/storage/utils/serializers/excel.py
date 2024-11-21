# -*- coding: utf-8 -*-

import collections
# @Author : Harold Chen
# @Date   : 2018/3/8
import copy
import logging
import os
import random
import re
import shutil
import string
from decimal import Decimal, InvalidOperation
from typing import Callable, Optional

from django.conf import settings
from django.db import transaction
from django.utils.translation import gettext
from mongoengine import DoesNotExist
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from apps.storage.docs.data import DataContent
from apps.storage.models.data import DataMetaFieldEnum, DataMeta, DataVisibility, DATA_VISIBILITY_NAME_MAP
from apps.storage.models.file import DataContentFile, DataContentImage
from apps.storage.models.material import MaterialSubject
from apps.storage.models.template import Template, TemplateField, TemplateFieldEnum
from apps.storage.utils.serializers.common import (FieldSerializer, FieldParser, ParsingError, ParsingErrorEnum,
                                                   DataMode, DataHandler)
from mgedata.errors.models import MGEError

logger = logging.getLogger('django')


def escape(value):
    if value is None:
        return
    value = str(value)
    CHAR_REGEX = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    def _sub(match):
        """
        Callback to escape chars
        """
        return ""

    return CHAR_REGEX.sub(_sub, value)


class ExcelFieldSerializer(FieldSerializer):
    """
    Excel字段序列化工具
    """

    def serialize_field(self, mongo_value, field_meta, meta_id):
        if mongo_value is None:
            return None
        t = field_meta.field_type
        if t.is_string:
            return str(mongo_value)
        elif t.is_number:
            return mongo_value
        elif t.is_range:
            field_type = field_meta.field_type
            if field_type.is_range_interval(field_meta.range_type):
                lb = mongo_value.get('lb', '-∞')
                ub = mongo_value.get('ub', '+∞')
                return f'({lb}, {ub})'
            else:
                err = mongo_value.get('err', '?')
                val = mongo_value.get('val', '?')
                return f'{val}±{err}'
        elif t.is_image or t.is_file:
            return self.serialize_image_file(mongo_value, meta_id)
        elif t.is_choice:
            return str(mongo_value)

    def serialize_image_file(self, mongo_value, meta_id):
        file_list = super(ExcelFieldSerializer, self).serialize_image_file(mongo_value, meta_id)
        if not file_list:
            return ""
        if len(file_list) > 1:
            return f'=HYPERLINK("{file_list[0]}", "{file_list[0]}, ......")'
        return f'=HYPERLINK("{file_list[0]}", "{file_list[0]}")'


class ExcelFieldParser(FieldParser):
    """
    Excel字段解析器
    """

    def parse_field(self, field_meta: TemplateField, raw_value):
        try:
            if raw_value is None:
                if field_meta.required:
                    raise ParsingError(ParsingErrorEnum.VALUE_MISSING, o1=field_meta.field_path_str)
                return None
            else:
                return super().parse_field(field_meta, str(raw_value))
        except ParsingError as e:
            if field_meta.parent_field is not None and field_meta.parent_field.field_type.is_generator \
                    and e.error == ParsingErrorEnum.VALUE_MISSING:
                return None
            raise e

    def parse_string(self, field_meta, raw_value: str):
        return raw_value

    def parse_number(self, field_meta, raw_value: str):
        try:
            return float(Decimal(raw_value))
        except (TypeError, ValueError, InvalidOperation):
            raise ParsingError(ParsingErrorEnum.INVALID_VALUE, o1=str(raw_value), o2=field_meta.field_type.description,
                               o3=field_meta.external_field_path_str)

    def parse_range(self, field_meta, raw_value: str):
        if TemplateFieldEnum.is_range_error(field_meta.range_type):
            try:
                err_data = raw_value.split('±')
                val = float(Decimal(err_data[0]))
                err = float(Decimal(err_data[1]))
                return {'val': val, 'err': err}
            except (IndexError, AttributeError, TypeError, ValueError, InvalidOperation):
                raise ParsingError(ParsingErrorEnum.INVALID_VALUE, o1=raw_value,
                                   o2=gettext("误差型"), o3=field_meta.external_field_path_str)
        else:
            try:
                trimmed = raw_value[1:-1]
                bounds = trimmed.split(',')
                return {'lb': float(Decimal(bounds[0])), 'ub': float(Decimal(bounds[1]))}
            except (IndexError, AttributeError, TypeError, ValueError, InvalidOperation):
                raise ParsingError(ParsingErrorEnum.INVALID_VALUE, o1=raw_value,
                                   o2=gettext("区间型"), o3=field_meta.external_field_path_str)

    def parse_file_image(self, field_meta, raw_value: str):
        path_list = re.split("['\n',',','，']", raw_value)
        url_list = []
        for path in path_list:
            if path is None or path.strip() == '':
                raise ParsingError(ParsingErrorEnum.VALUE_MISSING, o1=field_meta.field_path_str)
            url = self.parse_single_file_or_image(field_meta, path.strip())
            if url is not None:
                url_list.append(url)
        if len(url_list) == 0:
            if field_meta.required:
                raise ParsingError(ParsingErrorEnum.VALUE_MISSING, o1=field_meta.field_path_str)
        elif len(url_list) > 1 and not field_meta.has_multiple_files:
            raise ParsingError(ParsingErrorEnum.EXCESSIVE_FILES_OR_IMAGES, o1=field_meta.field_path_str)
        return url_list

    def parse_choice(self, field_meta, raw_value: str):
        choices = field_meta.choices_list
        if raw_value not in choices:
            raise ParsingError(ParsingErrorEnum.INVALID_CHOICE_VALUE, o1=raw_value,
                               o2=', '.join(choices))
        return raw_value


try:
    Mapping = collections.Mapping
except AttributeError:
    Mapping = collections.abc.Mapping


class Cursor(Mapping):
    """
    游标，记录某个sheet当前处理到哪个单元格
    """

    DATA_START_ROW_NUM = 2  # 数据行开始于第几行

    def __init__(self, row=DATA_START_ROW_NUM - 1, column=1):
        self.row = row
        self.column = column

    def move_left(self):
        """
        向左移动一个单元格
        :return: self
        """
        self.column -= 1
        if self.column == 0:
            self.column = 1
        return self

    def move_right(self):
        """
        向右一个单元格
        :return:
        """
        self.column += 1
        return self

    def move_down(self):
        """
        向下移动一个单元格
        :return:
        """
        self.column += 1
        return self

    def move_to_new_line(self, row=None, col_start_at=2):
        """
        移动到新的行，并且col_start_at对应的列，如果未指定row则移动到下一行
        :param row: 新的行号
        :param col_start_at: 列
        :return: self
        """
        if row:
            self.row = row
        else:
            self.row += 1
        self.column = col_start_at
        return self

    def move_to_line_head(self, col_start_at=2):
        """
        移动到当前行的第一列的单元格，默认为2
        :param col_start_at: 列
        :return: self
        """
        self.column = col_start_at
        return self

    def get_line_head(self, col_start_at=1):
        """
        返回一个新的游标，并置于本行第一个单元格
        :param col_start_at:
        :return:
        """
        return Cursor(row=self.row, column=col_start_at)

    def to_coordinate(self):
        """
        游标位置转为Excel字母+数字的坐标形式，例如A3单元格
        :return:
        """
        return get_column_letter(self.column) + str(self.row)

    """
    下列函数用于方便使用openpyxl的cell函数（**解包坐标）
    """

    def __getitem__(self, item):
        if item == 'row':
            return self.row
        elif item == 'column':
            return self.column

    def __len__(self):
        return 2

    def __iter__(self):
        return iter({'row': self.row, 'column': self.column})


# 字段坐标
class FieldCoordinate:
    def __init__(self, sheet_name, row, column, at_start, at_end, in_array):
        """
        :param sheet_name: sheet名称
        :param row: header所在行
        :param column: 列
        :param at_start: 是否是某行第一个数据字段（除去Data ID和Table ID）
        :param at_end: 是否是某行最后一个字段
        :param in_array: 是否在数组中
        """
        self.sheet_name = sheet_name
        self.row = row
        self.column = column
        self.at_start = at_start
        self.at_end = at_end
        self.in_array = in_array


class TableRowCursor:
    """
    如果Table类型字段嵌在数组中，对应的Sheet需要用这个Cursor来遍历，
    因为同一个DataID的Row可能属于不同的Table（不同的数组元素）
    """

    def __init__(self, start_row=Cursor.DATA_START_ROW_NUM - 1, table_id_column=2):
        self._row = start_row
        self._column = table_id_column
        self.current_table_id = None

    @property
    def row(self):
        return self._row

    def has_next_row(self, sheet, data_id, array_index, ignore_table_id=False):
        """
        判断下一行是否属于data_id对应的数据中的表格（并且是当前表格）中的行
        :param sheet: sheet
        :param data_id: 当前数据ID
        :param ignore_table_id: 无视Table ID，强制读取所有属于data_id的行
        :return: True or False
        """
        self._row += 1
        target_data_id = sheet.cell(row=self._row, column=1).value
        if target_data_id is not None:
            target_data_id = str(target_data_id)
        if target_data_id == data_id:
            if ignore_table_id:
                return True
            table_id = sheet.cell(row=self._row, column=self._column).value
            if self.current_table_id is None and table_id is not None:
                if table_id == array_index:
                    self.current_table_id = table_id
                    return True
            elif table_id is not None and table_id == self.current_table_id:
                return True
        self._row -= 1
        self.current_table_id = None
        return False


class ExcelHandler(DataHandler):
    """
    通用函数
    """

    # 数据元信息sheet名称
    @property
    def _meta_sheet_name(self):
        return gettext('元数据')

    # 第一级数据sheet名称
    @property
    def _root_data_name(self):
        return "1"

    def __init__(self, mode, template: Template = None, no_attachments=False, verify_only=False):
        """
        初始化
        :param mode: 操作模式
        :param template: 数据模板，如果是write或者template模式则必须提供
        """
        super().__init__(mode, template, no_attachments, verify_only=verify_only)
        self._meta_cursor = Cursor(row=1, column=1)  # 数据元信息Sheet游标
        self._reference_cursor = Cursor(row=1, column=1)
        self._reference_cursor_map = {}  # 记录field中需要隐藏处理的数据
        self._parser = None
        self._sheet_name_cursor_map = {}
        self._path_str_coordinate_map = {}
        self._path_str_sheet_name_map = {}
        self._path_str_table_cursor_map = {}
        self._path_str_sub_sheet_name_map = {}  # 记录表格型和数组型元素所在的sheet名称
        self._sheet_name_data_id_list_map = {}  # 记录每个sheet中每个DataID的所有行号
        self._sheet_name_path_str_map_list = []  # 记录sheet名称和字段的对应关系（创建目录使用）
        self._wb = None
        self._serializer = ExcelFieldSerializer()
        if self._mode == DataMode.WRITE or self._mode == DataMode.TEMPLATE:
            if template is None:
                raise ValueError(f'A template must be provided in {self._mode.description} mode.')
            self._load_template(template)

    def _get_cursor_by_field(self, field_meta: TemplateField):
        """
        通过字段path获取对应的sheet的cursor
        :param field_meta: 字段meta
        :return:
        """
        coordinate = self._get_field_coordinate_by_field(field_meta)
        return self._sheet_name_cursor_map[coordinate.sheet_name]

    def _get_field_coordinate_by_field(self, field_meta: TemplateField):
        """
        通过字段path获取其header的坐标
        :param field_meta: 字段meta
        :return:
        """
        return self._path_str_coordinate_map[field_meta.field_path_str]

    def _get_sheet_by_field(self, field_meta: TemplateField):
        """
        通过字段path获取对应的sheet
        :param field_meta: 字段meta
        :return:
        """
        coordinate = self._get_field_coordinate_by_field(field_meta)
        return self._wb[coordinate.sheet_name]

    def _get_sheet_name_by_field(self, field_meta: TemplateField):
        return self._get_field_coordinate_by_field(field_meta).sheet_name

    def _create_sheet(self, title=None, field_meta: TemplateField = None):
        """
        建一个新的sheet
        :param title: sheet名称，如果不提供则使用sheet的编号作为sheet名称
        :param field_meta: 字段meta
        :return:
        """
        ws = self._wb.create_sheet()
        if title is None or title == self._root_data_name:
            ws.title = str(self._wb.index(ws))
            ws.cell(row=Cursor.DATA_START_ROW_NUM - 1, column=1).value = gettext('数据ID')
        else:
            ws.cell(row=1, column=1).value = gettext('数据ID')
        if field_meta is not None:
            self._sheet_name_path_str_map_list.append((ws.title, field_meta.field_path_str))
        self._sheet_name_cursor_map[ws.title] = Cursor()

        return ws

    def _get_table_cursor_by_field(self, field_meta: TemplateField):
        return self._path_str_table_cursor_map[field_meta.field_path_str]

    def _record_reference_table_from_field(self, ws, column, field_meta: TemplateField):
        """
        记录filed_meta相关信息，延迟处理
        """
        field_path_str = field_meta.field_path_str
        cursor_map = self._reference_cursor_map.setdefault(ws.title, {})
        cursor_map[field_path_str] = dict(field_meta=field_meta, field_column=column)

    def _write_reference_table_from_field(self, ws, column, field_meta: TemplateField) -> Cursor:
        """
        在ws的指定column生成field_meta的值并隐藏
        :return: Cursor
        """
        reference_cursor = Cursor(row=1, column=1)
        for _ in range(column - 1):
            reference_cursor.move_right()
        for value in field_meta.choices_list:
            ws.cell(**reference_cursor).value = value
            ws.cell(**reference_cursor).number_format = ';;;'  # 隐藏单元格
            reference_cursor.move_to_new_line(col_start_at=column)
        return reference_cursor

    def _create_headers_from_field(self, sheet, field_meta: TemplateField, fill_header=True,
                                   at_start=False, at_end=False, in_array=False):
        """
        根据字段填写表头
        :param sheet: 需要填写的sheet
        :param field_meta: 字段meta
        :param fill_header: 是否填写该字段
        :param at_start: 该字段是否为sheet的第一个数据字段（除了Data ID和Table ID等）
        :param at_end: 该字段是否为sheet的最后一个字段
        :return: None
        """
        t = field_meta.field_type
        cursor = self._sheet_name_cursor_map[sheet.title]  # 获取sheet的游标
        if not (t.is_container or t.is_generator):
            cursor.move_right()

        if fill_header:  # 不是数组的直接元素
            if not (t.is_container or t.is_generator):
                unit = field_meta.unit
                header = field_meta.external_field_path_str
                if unit:
                    sheet.cell(**cursor).value = f'{header} ({unit})'  # 填写表头
                else:
                    sheet.cell(**cursor).value = header

        # 记录表头坐标
        self._path_str_coordinate_map[field_meta.field_path_str] = FieldCoordinate(sheet.title, cursor.row,
                                                                                   cursor.column, at_start, at_end,
                                                                                   in_array)

        if t.is_table:
            ws = self._create_sheet(field_meta=field_meta)  # 为表格建一个sheet
            self._path_str_sub_sheet_name_map[field_meta.field_name] = ws.title
            if in_array:
                ws.cell(row=Cursor.DATA_START_ROW_NUM - 1, column=2).value = gettext(
                    '元素序号')  # sheet的第二列是Item Index
                self._sheet_name_cursor_map[ws.title].move_right()

            self._path_str_table_cursor_map[field_meta.field_path_str] = TableRowCursor()
            sub_field_names = field_meta.sub_field_names
            for sub_field_meta in field_meta.sub_fields:  # 填写表格的每一列
                self._create_headers_from_field(ws, field_meta=sub_field_meta,
                                                at_start=sub_field_meta.field_name == sub_field_names[0],
                                                at_end=sub_field_meta.field_name == sub_field_names[-1],
                                                in_array=in_array)

        elif t.is_container or t.is_generator:
            ws = sheet
            sub_field_names = field_meta.sub_field_names
            for sub_field_meta in field_meta.sub_fields:
                self._create_headers_from_field(ws, field_meta=sub_field_meta,
                                                at_start=sub_field_meta.field_name == sub_field_names[0] and at_start,
                                                at_end=sub_field_meta.field_name == sub_field_names[-1] and at_end,
                                                in_array=in_array)

        elif t.is_array:
            ws = self._create_sheet(field_meta=field_meta)  # 为数组建一个sheet
            # self._path_str_sub_sheet_name_map[path_str] = ws.title
            element_meta = field_meta.element_field
            element_t = element_meta.field_type
            ws.cell(row=Cursor.DATA_START_ROW_NUM - 1, column=2).value = gettext('元素序号')  # 第二列是Item Index
            if not (element_t.is_container or element_t.is_generator):
                ws.cell(row=Cursor.DATA_START_ROW_NUM - 1, column=3).value = gettext('元素')  # 第三列是Items
            self._sheet_name_cursor_map[ws.title].move_right()

            self._create_headers_from_field(ws, field_meta=element_meta, fill_header=False, at_start=True,
                                            at_end=True,
                                            in_array=True,
                                            )

        elif t.is_choice:
            if self._mode == DataMode.TEMPLATE:
                self._record_reference_table_from_field(sheet, cursor.column, field_meta)
        if at_end and sheet.title in self._reference_cursor_map:
            hidden_column = cursor.column + 1
            for field_path_str, field_dict in self._reference_cursor_map[sheet.title].items():
                field_meta = field_dict['field_meta']
                reference_cursor = self._write_reference_table_from_field(sheet, hidden_column, field_meta)
                dv = DataValidation(type="list",
                                    formula1="=${}${}:${}${}".format(
                                        get_column_letter(hidden_column),
                                        1,
                                        get_column_letter(hidden_column),
                                        reference_cursor.row - 1
                                    ))
                column_letter = get_column_letter(field_dict['field_column'])
                dv.ranges.add(f'{column_letter}2:{column_letter}1048576')
                sheet.add_data_validation(dv)
                field_dict['hidden_column'] = hidden_column
                hidden_column += 1

    def _write_reference_table(self, ws, column):
        # TODO 只返回用户参与的项目和课题
        subjects = MaterialSubject.objects.all().select_related('project')
        project_name_subjects_map = {}

        for subject in subjects:
            project_name = subject.project.name
            project_id = subject.project.id
            subject_name = subject.name
            subject_id = subject.id
            key = f'{project_name}（{project_id}）'
            value = f'{subject_name}（{subject_id}）'
            project_name_subjects_map.setdefault(key, []).append(value)

        for _ in range(column):
            self._reference_cursor.move_right()
        for project, subject_list in project_name_subjects_map.items():
            for subject in subject_list:
                ws.cell(**self._reference_cursor).value = f'{project}/{subject}'
                ws.cell(**self._reference_cursor).number_format = ';;;'  # 隐藏单元格
                self._reference_cursor.move_to_new_line(col_start_at=column + 1)

    def _create_meta_sheet(self):
        """
        创建元数据sheet
        :return:
        """
        ws = self._wb.create_sheet(self._meta_sheet_name)
        if self._mode != DataMode.WRITE:
            self._write_reference_table(ws, len(DataMetaFieldEnum.reader_meta_fields()) + 2 + 3 + 1 + 1)
        ws['A1'] = gettext('数据ID')
        cursor = self._meta_cursor
        if self._mode == DataMode.TEMPLATE:  # 将template id记录在sheet中并隐藏
            template_id_cell = ws.cell(row=cursor.row, column=len(DataMetaFieldEnum.reader_meta_fields()) + 2 + 3 + 1)
            template_id_cell.value = str(self._template.id) + gettext(' （请不要修改或删除此单元格的内容！！！！！！！）')
            template_id_cell.number_format = ';;;'  # 隐藏单元格
            iterator = DataMetaFieldEnum.reader_iterator()
        else:
            iterator = DataMetaFieldEnum.writer_iterator()
        for index, field in enumerate(iterator):
            if field == DataMetaFieldEnum.PROJECT:
                continue
            cursor.move_right()
            # 项目课题合为一列
            if field == DataMetaFieldEnum.SUBJECT:
                header = '机构'
            else:
                header = field.description
            ws.cell(row=cursor.row, column=cursor.column).value = header

            if field == DataMetaFieldEnum.SUBJECT and self._mode == DataMode.TEMPLATE:
                dv = DataValidation(type="list",
                                    formula1="=${}${}:${}${}".format(
                                        get_column_letter(self._reference_cursor.column),
                                        1,
                                        get_column_letter(self._reference_cursor.column),
                                        self._reference_cursor.row - 1
                                    ))
                column_letter = get_column_letter(cursor.column)
                dv.ranges.add(f'{column_letter}2:{column_letter}1048576')
                ws.add_data_validation(dv)

            if field == DataMetaFieldEnum.VISIBILITY and self._mode == DataMode.TEMPLATE:
                dv = DataValidation(type="list",
                                    formula1='"' + ",".join([option.description for option in
                                                             DataVisibility]) + '"')
                column_letter = get_column_letter(cursor.column)
                dv.ranges.add(f'{column_letter}2:{column_letter}1048576')
                ws.add_data_validation(dv)

        cursor.move_to_new_line(col_start_at=1)

    def _create_map_sheet(self):
        """
        建立目录sheet
        :return:
        """
        if len(self._sheet_name_path_str_map_list) == 0:
            return

        ws = self._wb.create_sheet(gettext('目录'), index=0)
        ws['A1'].value = gettext('工作表名')
        ws['B1'].value = gettext('内容')
        ws['A2'].value = '1'
        text = gettext("第一级数据字段")
        ws['B2'].value = f'=HYPERLINK("#1!A2", "{text}")'
        ws['B2'].style = 'Hyperlink'
        row = 3

        for sheet_name, path_str in self._sheet_name_path_str_map_list:
            detail = gettext("'%(field_name)s'的内容") % {'field_name': path_str}
            ws.cell(row=row, column=1).value = sheet_name
            cell = ws.cell(row=row, column=2)
            cell.value = f'=HYPERLINK("#{sheet_name}!A1", "{detail}")'
            cell.style = 'Hyperlink'
            row += 1

    def _load_template(self, template: Template):
        self._template = template
        self._wb = Workbook()  # 新建一个工作簿（即使是read模式，也要先建立一个以记录字段的位置）
        self._create_meta_sheet()
        self._wb.remove(self._wb.active)  # 删除自带的第一个sheet
        ws = self._create_sheet(self._root_data_name)
        field_names = template.field_names
        fields = template.fields
        start = field_names[0]
        end = field_names[-1]
        for field in fields:
            self._create_headers_from_field(ws, field_meta=field,
                                            at_start=field.field_name == start, at_end=field.field_name == end)
        [x.move_to_new_line(row=Cursor.DATA_START_ROW_NUM, col_start_at=2) for x in
         self._sheet_name_cursor_map.values()]

    def _adjust_format(self):
        """
        调整列宽、居中的样式
        :return:
        """
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for sheet_name in self._wb.sheetnames:
            sheet = self._wb[sheet_name]
            if sheet_name == self._meta_sheet_name:
                for index, col in enumerate(sheet.columns):
                    if self._mode == DataMode.TEMPLATE and index == len(DataMetaFieldEnum.reader_meta_fields()) + 1:
                        continue
                    column_name = get_column_letter(col[0].column)
                    max_width = 0
                    for cell in col:
                        try:  # 防止单元格没有内容
                            cell.alignment = alignment
                            cell_width = len(cell.value.encode('utf-8'))
                            if cell_width > max_width:
                                max_width = cell_width
                        except (ValueError, TypeError, AttributeError):
                            pass
                    adjusted_width = max_width
                    sheet.column_dimensions[column_name].width = adjusted_width * 1.15
            else:
                for col in sheet.columns:
                    column_name = get_column_letter(col[0].column)
                    max_width = 0
                    for cell in col:
                        try:  # 防止单元格没有内容
                            cell.alignment = alignment
                            links = str(cell.value).split('\n')
                            if len(links) > 0:
                                cell_width = len(links[0].encode('utf-8'))
                            else:
                                cell_width = len(str(cell.value).encode('utf-8'))
                            if cell_width > max_width:
                                max_width = cell_width
                        except (ValueError, TypeError, IndexError):
                            pass
                    adjusted_width = max_width
                    sheet.column_dimensions[column_name].width = adjusted_width * 1.15

    def save(self, output_dir):
        """
        保存excel
        :param output_dir: 导出到文件夹
        :return: 导出的excel文件名
        """
        self._create_map_sheet()
        self._adjust_format()
        if not self._no_attachments:
            for src, new_filename in self._serializer.filenames:
                try:
                    shutil.copy(src, os.path.join(output_dir, new_filename))
                except FileNotFoundError:
                    logger.error(f"附件丢失：{src}")
                    continue
        excel_filename = f'{self._template.title.replace("/", "")}.xlsx'
        self._wb.save(os.path.join(output_dir, excel_filename))
        return excel_filename

    """
    导出专用函数
    """

    def write_data(self, data_meta: DataMeta):
        """
        添加一个数据
        :param data_meta: 数据的元数据
        :return: None
        """
        if self._mode != DataMode.WRITE:
            raise ValueError(f"Cannot write excel in {self._mode.description} mode.")
        if not isinstance(data_meta, DataMeta):
            raise ValueError("data_meta should be an instance of DataMeta")
        if data_meta.template_id != self._template.id:
            raise ValueError(f'Data\'s tid should be "{self._template.id}", got "{data_meta.template_id}" instead')

        data_id = str(data_meta.id)

        ws = self._wb[self._meta_sheet_name]
        ws.cell(**self._meta_cursor).value = data_id  # 元数据sheet第一列是Data ID
        self._meta_cursor.move_right()
        for index, field in enumerate(DataMetaFieldEnum.writer_iterator()):
            # 填写其他元数据
            if field == DataMetaFieldEnum.PROJECT:
                continue
            elif field == DataMetaFieldEnum.SUBJECT:
                value = f"{data_meta.project.name}({data_meta.project_id})/{data_meta.subject.name}({data_meta.subject_id})"
                ws.cell(**self._meta_cursor).value = value
            else:
                try:
                    ws.cell(**self._meta_cursor).value = escape(field.get_value(data_meta))
                except Exception:
                    logging.error(f"Value that caused the exception: {field.get_value(data_meta)}")
                    raise
            self._meta_cursor.move_right()
        self._meta_cursor.move_to_new_line(col_start_at=1)

        def _add_field(field_meta: TemplateField, field_data, fillings_at_start=(data_id,),
                       force_new_line_when_done=False) -> str:
            """
            填写一个字段的数据

            :param field_data: 字段的原始数据
            :param fillings_at_start: 开始若干列手动填充的数据（例如Data ID，Table ID）
            :param force_new_line_when_done: 填写完该字段后是否强制换行
            :return: 链接到cell所在行首个cell的超链接字符串
            """
            t = field_meta.field_type
            value = None  # 对应cell填写的内容，初始为空，待解析
            sheet = self._get_sheet_by_field(field_meta)  # 找到该字段对应的sheet
            coordinate = self._get_field_coordinate_by_field(field_meta)  # 找到该字段表头的坐标
            cursor = self._get_cursor_by_field(field_meta)  # 找到sheet的cursor
            cell = sheet.cell(row=cursor.row, column=coordinate.column)  # 需要填写的cell

            if field_data is not None:
                if t.is_table:
                    hyperlinks = []
                    for row in field_data:
                        for header in field_meta.sub_fields:
                            hyperlink = _add_field(header, row.get(header.field_name),
                                                   fillings_at_start=fillings_at_start)
                            hyperlinks.append(hyperlink)
                    if len(hyperlinks):
                        # 只需要第一个元素超链接
                        value = '=HYPERLINK("{}", "{}")'.format(hyperlinks[0], gettext("点此查看"))
                        cell.style = 'Hyperlink'
                elif t.is_container or t.is_generator:
                    if field_data:
                        for sub_field in field_meta.sub_fields:
                            _add_field(sub_field, field_data.get(sub_field.field_name),
                                       fillings_at_start=fillings_at_start)
                elif t.is_array:
                    hyperlinks = []
                    for i, element in enumerate(field_data):
                        # 数组元素除了填充Data ID还需要填充Item Index
                        hyperlinks.append(
                            _add_field(field_meta.element_field, field_data=element,
                                       fillings_at_start=fillings_at_start + (str(i + 1),)))
                    if len(hyperlinks):
                        value = '=HYPERLINK("{}", "{}")'.format(hyperlinks[0], gettext("点此查看"))
                        cell.style = 'Hyperlink'
                else:  # 基本类型
                    value = self._serializer.serialize_field(field_data, field_meta, data_meta.id)
                if value is not None:
                    cell.value = escape(value)
                    if field_meta.field_type.is_file or field_meta.field_type.is_image:
                        cell.style = 'Hyperlink'
            else:
                # 如果field_data为空，仍然继续递归，防止后续代码不执行导致不换行
                if t.is_table or t.is_generator or t.is_container:
                    for f in field_meta.sub_fields:
                        _add_field(f, None, fillings_at_start)
                elif t.is_array:
                    _add_field(field_meta.element_field, None, fillings_at_start)
            # 如果是第一个数据字段，同时填充其左侧的Data ID及其他字段
            if not (t.is_container or t.is_generator) and coordinate.at_start:
                t_cursor = cursor.get_line_head()
                for filling in fillings_at_start:
                    sheet.cell(**t_cursor).value = filling
                    t_cursor.move_right()
            line_head = cursor.get_line_head()
            # 如果是最后一个字段，或者强制换行（生成器的字段）时，换行
            if not (t.is_container or t.is_generator) and (coordinate.at_end or force_new_line_when_done):
                cursor.move_to_new_line()
            # 返回超链接，链接到本行第一个cell
            hyperlink = '#{}!{}'.format(sheet.title, line_head.to_coordinate())
            return hyperlink

        try:
            content = DataContent.objects.get(pk=data_meta.dc_id).data
        except DoesNotExist:
            content = {}  # TODO 找不到
            logger.error(f"找不到DataContent (meta_id={data_id}, dc_id={data_meta.dc_id})")

        for field in self._template.fields:
            _add_field(field, content.get(field.field_name))

            """
            导入专用函数
            """

    def _get_field_row_list(self, field_meta: TemplateField, data_id):
        return self._sheet_name_data_id_list_map[self._get_sheet_name_by_field(field_meta)].get(
            data_id)

    def read_data(self, fp_or_path, uploaded_by, file_dir=None,
                  progress_handler: Callable[[float, Optional[bool]], None] = None):
        """
        用excel中读取数据
        :param fp_or_path: excel文件路径或者fp
        :param uploaded_by: 上传的用户名
        :param file_dir: 数据中图片和文件所在的文件夹
        :param progress_handler: 更新进度时调用的回调函数
        :return: DataMeta ID列表
        """
        if self._mode != DataMode.READ:
            raise ValueError("Cannot read excel in %s mode." % (self._mode.description,))

        data_id = 0  # 当前正在解析的数据ID
        current_sheet_name = ""  # 当前正在解析的sheet名称
        current_row_num = 0  # 当前正在解析的行号
        current_column_num = 0  # 当前正在解析的列号
        if self._parser is None:
            self._parser = ExcelFieldParser(uploaded_by, file_dir, verify_only=self._verify_only)
        parser = self._parser
        current_count = 0
        total_count = 0

        def _read_field(field_meta: TemplateField, at_row, array_index=0):
            """
            读取某个字段的值
            :param field_meta: 字段meta
            :param at_row: 要读取的值所在的行号
            :param array_index: 读取的值在数组中的序号（如果是数组的元素，或者是内嵌在数组元素中），下标从1开始
            :return: 字段的json值
            """
            field_coordinate = self._get_field_coordinate_by_field(field_meta)
            current_sheet = self._get_sheet_by_field(field_meta)
            nonlocal current_sheet_name, current_column_num
            current_column_num = field_coordinate.column
            current_sheet_name = current_sheet.title
            t = field_meta.field_type
            required = field_meta.required

            if t.is_container or t.is_generator:
                json_data = {}
                # 取出容器子字段的row_list(所有子字段都在同一个sheet所以可以用第一个字段_sub_ord[0]取）
                row_list = self._get_field_row_list(field_meta[0], data_id)
                if row_list is None:
                    if required:
                        raise ParsingError(ParsingErrorEnum.VALUE_MISSING, o1=field_meta.external_field_path_str)
                    return None
                for sub_field in field_meta.sub_fields:
                    if array_index:  # 数组中的第array_index个元素
                        read_row_at = row_list[array_index - 1]
                    else:
                        read_row_at = row_list[0]

                    d = _read_field(sub_field, at_row=read_row_at, array_index=array_index)
                    if d:
                        json_data[sub_field.field_name] = d
                        if t.is_generator:
                            generator_choice_field = copy.deepcopy(sub_field)
                            generator_choice_field._parent_field = None
                            _read_field(generator_choice_field, at_row=read_row_at, array_index=array_index)
                            break  # 生成器只认第一个读到的字段
                if not json_data:
                    return None
            elif t.is_table:
                sub_sheet = self._get_sheet_by_field(field_meta[0])
                table_cursor = self._get_table_cursor_by_field(field_meta)
                json_data = []
                while table_cursor.has_next_row(sub_sheet, data_id, array_index=array_index,
                                                ignore_table_id=not field_coordinate.in_array):
                    json_data.append({})
                    for header in field_meta.sub_fields:
                        json_data[-1][header.field_name] = _read_field(header, table_cursor.row)
                    if not json_data[-1]:
                        json_data.pop()
                if len(json_data) == 0:
                    return None
            elif t.is_array:
                row_list = self._get_field_row_list(field_meta.element_field, data_id)
                if row_list is None:
                    if required:
                        raise ParsingError(ParsingErrorEnum.VALUE_MISSING, o1=field_meta.external_field_path_str)
                    return None
                json_data = []
                for index, row_number in enumerate(row_list):
                    element = _read_field(field_meta.element_field, row_number, array_index=index + 1)
                    if element is not None:
                        json_data.append(element)
                if len(json_data) == 0:
                    if required:
                        raise ParsingError(ParsingErrorEnum.NO_ELEMENTS_IN_ARRAY,
                                           o1=field_meta.external_field_path_str)
                    else:
                        return None
            else:
                if required and current_sheet.cell(row=at_row, column=field_coordinate.column).value in (None, [None], [], [""]):
                    raise ParsingError(ParsingErrorEnum.VALUE_MISSING, o1=field_meta.external_field_path_str)
                    return None
                json_data = parser.parse_field(field_meta,
                                               current_sheet.cell(row=at_row, column=field_coordinate.column).value)
            return json_data

        def _read_meta():
            nonlocal current_sheet_name, current_row_num, current_column_num
            current_sheet_name = self._meta_sheet_name
            try:
                ws = self._wb[self._meta_sheet_name]
            except KeyError:
                raise MGEError.BAD_TEMPLATE(f"缺少工作表\"{self._meta_sheet_name}\"")
            did_meta_map = {}
            for meta_row in ws.iter_rows(min_row=2):
                did = meta_row[0].value
                if did is None or isinstance(did, str) and did.strip() == "":
                    break
                did = str(did)
                if did in did_meta_map:
                    continue
                current_row_num = meta_row[0].row
                temp_meta = {'tid': template_id}
                has_column = False
                base = 1
                read_fields = list(DataMetaFieldEnum.reader_meta_fields())
                # project和subject合并，只读取subject
                read_fields.remove(DataMetaFieldEnum.PROJECT)
                for index, field in enumerate(read_fields):
                    current_column_num = index + base + 1
                    value = meta_row[index + base].value
                    if field == DataMetaFieldEnum.SUBJECT:
                        value = meta_row[index + base].value
                        # 项目名(项目号)/课题名(课题号)
                        # 项目号和课题号可能包含字母
                        if value is None:
                            raise ParsingError(ParsingErrorEnum.VALUE_MISSING, o1=gettext("机构"))
                        value = str(value)
                        regex = re.compile('(^.+)[（(](.+)[）)]/(.+)[（(](.*)[）)]$')
                        match = regex.match(value)
                        if match is None:
                            raise ParsingError(ParsingErrorEnum.MALFORMED_PROJECT_SUBJECT)
                        project_id = match.group(2)
                        subject_id = match.group(4)
                        temp_meta[DataMetaFieldEnum.PROJECT.field_raw_name] = DataMetaFieldEnum.PROJECT.validate(
                            project_id)
                        temp_meta[DataMetaFieldEnum.SUBJECT.field_raw_name] = DataMetaFieldEnum.SUBJECT.validate(
                            subject_id)

                    elif field == DataMetaFieldEnum.VISIBILITY:
                        if value is None:
                            raise ParsingError(ParsingErrorEnum.VALUE_MISSING, o1="公开范围")
                        visibility = DATA_VISIBILITY_NAME_MAP.get(str(value))
                        if visibility is None:
                            raise ParsingError(
                                ParsingErrorEnum.INVALID_CHOICE_VALUE,
                                o1=str(value),
                                o2='，'.join(
                                    [x.description for x in DataVisibility]
                                )
                            )
                        temp_meta[field.field_raw_name] = visibility
                    else:
                        if value is None:
                            value = field.validate(None)
                        else:
                            value = field.validate(str(value))
                        if value is not None:
                            temp_meta[field.field_raw_name] = value
                            has_column = True

                if has_column:
                    did_meta_map[str(meta_row[0].value)] = temp_meta
            did_meta_map.pop(None, None)
            return did_meta_map

        def _read_data():
            nonlocal current_count, total_count
            dd = {}
            for row in self._wb[self._root_data_name].iter_rows(min_row=Cursor.DATA_START_ROW_NUM):
                nonlocal current_row_num, data_id
                current_row_num = row[0].row
                data_id = row[0].value
                if data_id is None:
                    continue
                data_id = str(data_id)
                if data_id in dd:
                    continue
                data = {}
                for field in template.fields:
                    data[field.field_name] = _read_field(field, at_row=row[0].row)
                    dd[data_id] = data

                current_count += 1
                if progress_handler:
                    progress_handler(current_count / total_count)
            return dd

        try:
            try:
                if progress_handler:
                    progress_handler(-1, True)
                wb = load_workbook(fp_or_path, data_only=True)
            except Exception:
                raise MGEError.BAD_EXCEL
            try:
                meta_sheet = wb[self._meta_sheet_name]
            except KeyError:
                raise MGEError.BAD_TEMPLATE(f"缺少工作表\"{self._meta_sheet_name}\"")
            template_id = meta_sheet.cell(row=1, column=len(DataMetaFieldEnum.reader_meta_fields()) + 5 + 1).value
            try:
                r = re.search(r'^[1-9]\d*', str(template_id))
                template_id = r.group()
                if template_id == '':
                    raise TypeError()
                template = Template.objects.get(id=template_id)
                self._load_template(template)
            except (TypeError, AttributeError):
                raise ParsingError(ParsingErrorEnum.BAD_TEMPLATE)
            except Template.DoesNotExist:
                raise ParsingError(ParsingErrorEnum.TEMPLATE_NOT_FOUND, o1=template_id)
            self._wb = wb
            # 记录每个sheet每个数据ID所在的行
            for sheet_name in self._wb.sheetnames:
                sheet = self._wb[sheet_name]
                self._sheet_name_data_id_list_map[sheet_name] = {}
                for row in sheet.iter_rows(min_col=1, max_col=1, min_row=2):
                    data_id = row[0].value
                    if data_id is not None:
                        data_id = str(data_id)
                    else:
                        continue
                    data_row_number_list = self._sheet_name_data_id_list_map[sheet_name].get(data_id)
                    if data_row_number_list is None:
                        data_row_number_list = []
                        self._sheet_name_data_id_list_map[sheet_name][data_id] = data_row_number_list
                    data_row_number_list.append(row[0].row)
            if progress_handler:
                progress_handler(0.002, True)
            meta_dict = _read_meta()  # 元数据，key为数据id
            if self._verify_only:
                total_count = len(meta_dict)
            else:
                total_count = len(meta_dict) * 2  # 认为数据解析和写数据库各占一半时间
            if total_count == 0:
                raise ParsingError(ParsingErrorEnum.FIRST_DATA_IS_BLANK)
            data_dict = _read_data()  # 数据，key为数据id
            if not self._verify_only:
                if progress_handler:
                    progress_handler(-1, True)
                data_images = []
                data_files = []
                temp_dir_name = "__tmp_dir_for_packing__"
                temp_dir = os.path.join(file_dir, temp_dir_name)
                os.makedirs(temp_dir, exist_ok=True)
                # 把图片和文件重命名，打包，移动到media，解包
                date_part = parser.date_prefix
                file_dst_dir = os.path.join(temp_dir, settings.DATA_FILE_PREFIX, date_part)
                image_dst_dir = os.path.join(temp_dir, settings.DATA_IMAGE_PREFIX, date_part)
                os.makedirs(file_dst_dir, exist_ok=True)
                os.makedirs(image_dst_dir, exist_ok=True)
                for image_path, packed in parser.file_image_path_info_map.items():
                    new_filename, hash_value, size = packed
                    data_images.append(
                        DataContentImage(file=os.path.join(settings.DATA_IMAGE_PREFIX, date_part, new_filename),
                                         hash_value=hash_value,
                                         size=size))
                    shutil.move(os.path.join(file_dir, image_path), os.path.join(image_dst_dir, new_filename))
                for file_path, packed in parser.file_path_info_map.items():
                    new_filename, hash_value, size = packed
                    data_files.append(
                        DataContentFile(file=os.path.join(settings.DATA_FILE_PREFIX, new_filename),
                                        hash_value=hash_value,
                                        size=size))
                    shutil.move(os.path.join(file_dir, file_path), os.path.join(file_dst_dir, new_filename))
                if data_files or data_images:
                    rand = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(10))
                    archive_name = rand
                    archive_path = os.path.join(file_dir, archive_name)
                    shutil.make_archive(archive_path, 'zip', temp_dir, settings.MEDIA_PREFIX)
                    archive_path += '.zip'
                    archive_name += '.zip'
                    new_archive_path = os.path.join(settings.MEDIA_ROOT, archive_name)
                    shutil.move(archive_path, new_archive_path)
                    shutil.unpack_archive(new_archive_path, settings.MEDIA_ROOT, 'zip')
                    os.remove(new_archive_path)
                    parser.file_path_info_map = {}  # 清空， 防止read被再次调用导致再次移动文件
                    parser.file_image_path_info_map = {}
                with transaction.atomic():
                    DataContentFile.objects.bulk_create(data_files, ignore_conflicts=True)
                    DataContentImage.objects.bulk_create(data_images, ignore_conflicts=True)

            for data_id in data_dict:
                try:
                    meta = meta_dict[data_id]
                except KeyError:
                    raise ParsingError(ParsingErrorEnum.NO_META, no_position_info=True, o1=data_id)
                content = data_dict[data_id]
                if not self._verify_only:
                    # 如果是只验证，不执行下面这行
                    self.meta_id_list.append(
                        self._template.add_data(meta, content, uploaded_by=uploaded_by, importing=True).id)
                    current_count += 1
                    if progress_handler:
                        progress_handler(current_count / total_count)
            return self.meta_id_list
        except ParsingError as e:
            if not e.no_position_info and current_sheet_name and current_row_num and current_column_num:
                suffix = gettext('（工作表"%(sheet_name)s"中的单元格%(column_letter)s%(row)s）') % {
                    'sheet_name': current_sheet_name,
                    'column_letter': get_column_letter(
                        current_column_num), 'row': current_row_num}
                e.add_detail(suffix)
            if not self._verify_only:
                self.roll_back()
            raise e
        except Exception:
            if not self._verify_only:
                self.roll_back()
            raise

    """
    生成模板专用函数
    """

    def generate_template(self, number_of_data=3, elements_per_array=1, rows_per_table=1):
        if self._mode != DataMode.TEMPLATE:
            raise ValueError(f'Cannot generate template in {self._mode.description} mode')

        for data_count in range(1, number_of_data + 1):
            data_id = gettext('第%(number)s条数据的数据ID') % {'number': data_count}
            ws = self._wb[self._meta_sheet_name]
            ws.cell(**self._meta_cursor).value = data_id  # 第一列是Data ID
            self._meta_cursor.move_right()

            for index, field in enumerate(DataMetaFieldEnum.reader_iterator()):
                # 填写其他元数据
                if field == DataMetaFieldEnum.PROJECT:
                    continue
                ws.cell(**self._meta_cursor).value = gettext('在这里填写“%(field_name)s”') % {
                    'field_name': field.description}
                if field == DataMetaFieldEnum.SUBJECT:
                    ws.cell(**self._meta_cursor).value = gettext("请选择")
                elif field == DataMetaFieldEnum.VISIBILITY:
                    ws.cell(**self._meta_cursor).value = DataVisibility.PUBLIC.description
                else:
                    ws.cell(**self._meta_cursor).value = gettext('在这里填写“%(field_name)s”') % {
                        'field_name': field.description}
                self._meta_cursor.move_right()

            self._meta_cursor.move_to_new_line(col_start_at=1)

            def _add_field(field_meta: TemplateField, fillings_at_start=(data_id,)):
                t = field_meta.field_type
                sheet = self._get_sheet_by_field(field_meta)  # 找到该字段对应的sheet
                coordinate = self._get_field_coordinate_by_field(field_meta)  # 找到该字段表头的坐标
                cursor = self._get_cursor_by_field(field_meta)  # 找到sheet的cursor
                cell = sheet.cell(row=cursor.row, column=coordinate.column)  # 需要填写的cell
                value = ''
                if t.is_table:
                    for i in range(1, rows_per_table + 1):
                        for header in field_meta.sub_fields:
                            fillings = fillings_at_start + (
                                gettext(
                                    '此处输入表格的序号（例如2表示第2个表格）'),) if coordinate.in_array else fillings_at_start
                            _add_field(header, fillings_at_start=fillings)
                    sheet_name = self._get_sheet_name_by_field(field_meta[0])
                    text = gettext(
                        '=HYPERLINK("#%(sheet_name)s!A1", "请填写工作表 ""%(sheet_name)s""，此处不填")')
                    value = text % {
                        'sheet_name': sheet_name}
                    cell.style = 'Hyperlink'
                elif t.is_container or t.is_generator:
                    for sub_field in field_meta.sub_fields:
                        _add_field(sub_field)
                    return
                elif t.is_array:
                    for i in range(1, elements_per_array + 1):
                        _add_field(field_meta.element_field, fillings_at_start=fillings_at_start + ('',))
                    sheet_name = self._get_sheet_name_by_field(field_meta.element_field)
                    text = gettext(
                        '=HYPERLINK("#%(sheet_name)s!A1", "请前往工作表""%(sheet_name)s""填写")')
                    value = text % {
                        'sheet_name': sheet_name}
                    cell.style = 'Hyperlink'
                elif t.is_choice:
                    value = gettext('请从下拉框选择')
                else:  # 基本类型
                    example = ''
                    if t.is_range:
                        if t.is_range_error(field_meta.range_type):
                            example = gettext('（例如. a ± b）')
                        else:
                            example = gettext('（例如 (a, b) 且必须有括号）')
                    elif t.is_image or t.is_file:
                        example = gettext('（请填写文件或图片相对于本Excel文件的路径）')

                    value = gettext("在这里填写“%(field_name)s”（%(field_type)s）") % {
                        'field_name': field_meta.field_name,
                        'field_type': t.description
                    }
                    value += example

                cell.value = value
                # 如果是第一个数据字段，同时填充其左侧的Data ID及其他字段
                if coordinate.at_start:
                    t_cursor = cursor.get_line_head()
                    for filling in fillings_at_start:
                        sheet.cell(**t_cursor).value = filling
                        t_cursor.move_right()
                if coordinate.at_end:
                    cursor.move_to_new_line()

            for field in self._template.fields:
                _add_field(field)
