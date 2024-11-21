# -*- coding: utf-8 -*-

# @File   : flat_excel
# @Author : Harold Chen
# @Date   : 2018/4/16
import os
import shutil

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from apps.storage.models.data import DataMeta
from apps.storage.models.template import Template
from openpyxl.styles import Alignment
from django.utils.translation import gettext as _
from .flat_json import FlatJSONWriter
from .excel import ExcelFieldSerializer


class FlatExcelWriter:

    def __init__(self, template: Template):
        self._serializer = ExcelFieldSerializer()
        self._flat_json_writer = FlatJSONWriter(template, serializer=self._serializer, with_headers=True,
                                                with_meta=True)
        self._wb: Workbook = None
        self._template = template

    def write_data(self, data_meta: DataMeta):
        self._flat_json_writer.write_data(data_meta)

    def _adjust_format(self):
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet = self._wb.active
        for col in sheet.columns:
            column_name = col[0].column  # 获取列号
            column_name = get_column_letter(column_name)
            max_width = 0
            for cell in col:
                cell.alignment = alignment
                if cell.value is None:
                    continue
                links = cell.value.split('\n')
                if len(links) > 0:
                    cell_width = len(links[0].encode('utf-8'))
                else:
                    cell_width = len(cell.value.encode('utf-8'))
                if cell_width > max_width:
                    max_width = cell_width
            adjusted_width = max_width
            sheet.column_dimensions[column_name].width = adjusted_width * 1.2

    def save(self, output_dir):
        final_data = self._flat_json_writer.save()
        headers = [_("Index")] + final_data['data_headers']
        data_map_list = final_data['data']
        self._wb = Workbook()
        sheet = self._wb.active
        # sheet.title = 'Data'  # TODO 改名
        for index, header in enumerate(headers):
            sheet.cell(row=1, column=index + 1).value = header
        for data_index, data_array in enumerate(data_map_list):
            sheet.cell(row=data_index + 2, column=1).value = str(data_index + 1)
            for col_index, data_element in enumerate(data_array):
                if data_element is None:
                    continue
                sheet.cell(row=data_index + 2, column=col_index + 2).value = str(data_element)
        self._adjust_format()
        for src, new_filename in self._serializer.filenames:
            shutil.copy(src, os.path.join(output_dir, new_filename))
        excel_filename = f'{self._template.title.replace("/", "")}.xlsx'
        self._wb.save(os.path.join(output_dir, excel_filename))
        return excel_filename
