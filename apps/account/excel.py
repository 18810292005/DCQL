"""
批量创建用户
"""
import re
from enum import Enum

from django.db import transaction
from django.utils.translation import gettext as _
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet import worksheet

from apps.account.models.users import UserRole, User
from mgedata.errors.models import MGEError

EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')


class Header(Enum):
    USERNAME = 'username'
    PASSWORD = 'password'
    ROLE = 'role'
    REAL_NAME = 'real_name'
    EMAIL = 'email'
    INSTITUTION = 'institution'
    TELEPHONE = 'tel'

    @property
    def description(self):
        return {
            Header.USERNAME: _('用户名'),
            Header.PASSWORD: _('密码'),
            Header.ROLE: _('权限'),
            Header.REAL_NAME: _('真实姓名'),
            Header.EMAIL: _('邮箱'),
            Header.INSTITUTION: _('机构'),
            Header.TELEPHONE: _('电话')
        }[self]


# def generate_excel() -> bytes:
#
#     wb = Workbook()
#     ws: worksheet = wb.active
#
#     role_index = None
#     for i, header in enumerate(Header):
#         ws.cell(row=1, column=i + 1, value=header.description)
#         if header == Header.ROLE:
#             role_index = i
#     role_column_name = get_column_letter(role_index + 1)
#     role_choices = [str(x.description) for x in
#                     (UserRole.GUEST, UserRole.RESEARCHER, UserRole.DATA_ADMIN, UserRole.SYS_ADMIN)]
#     dv = DataValidation(type="list",
#                         formula1='"' + ",".join(role_choices) + '"')
#     dv.ranges.add(f'{role_column_name}2:{role_column_name}1048576')
#     ws.add_data_validation(dv)
#
#     alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     for index, col in enumerate(ws.columns):
#         column_name = get_column_letter(col[0].column)
#         ws.column_dimensions[column_name].width = 20
#     # align first 100 rows
#     for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=len(Header)):
#         for cell in row:
#             cell.alignment = alignment
#
#     in_mem_file = BytesIO()
#     wb.save(in_mem_file)
#     return in_mem_file.getvalue()
#

def _raise_value_missing(header: Header, cell: Cell):
    cell_name = cell.coordinate
    raise MGEError.INVALID_EXCEL_CELL_VALUE(
        _('列"%(header)s"的值不能为空（单元格%(cell_name)s）') % {
            'header': header.description, 'cell_name': cell_name
        }
    )


def _raise_invalid_value(msg: str, cell: Cell):
    cell_name = cell.coordinate
    raise MGEError.INVALID_EXCEL_CELL_VALUE(
        _('%(msg)s（错误位于单元格%(cell_name)s）') % {
            'cell_name': cell_name, 'msg': msg
        }
    )


def _check_none(value: str, header: Header, cell: Cell):
    if not value:
        _raise_value_missing(header, cell)


def read_excel(fp_or_path) -> list[User]:
    wb = load_workbook(fp_or_path, read_only=True)
    ws: worksheet = wb.active
    users = []
    username_row_map = {}
    email_row_map = {}
    if ws.max_row < 3:
        raise MGEError.INVALID_EXCEL_CELL_VALUE("Excel文件未填写")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if not any([cell.value for cell in row]):
            continue
        cur_user = User()
        users.append(cur_user)
        for i, header in enumerate(Header):
            cell: Cell = row[i]
            value = str(cell.value).strip() if cell.value is not None else None
            if header == Header.USERNAME:
                _check_none(value, header, cell)
                if value in username_row_map:
                    _raise_invalid_value(_('用户名与第%s行重复') % username_row_map[value], cell)
                if len(value) > 50:
                    _raise_invalid_value(_('用户名不能超过50个字符'), cell)
                username_row_map[value] = cell.row
                if User.objects.filter(username=value).exists():
                    _raise_invalid_value(_('系统中已存在用户名"%s"') % value, cell)
                cur_user.username = value
            elif header == Header.PASSWORD:
                _check_none(value, header, cell)
                cur_user.set_password(value)
            elif header == Header.ROLE:
                _check_none(value, header, cell)
                for role in UserRole:
                    if role.description == value:
                        cur_user.role = role
                        break
                else:
                    available_roles = [
                        str(role.description) for role in
                        (UserRole.GUEST, UserRole.RESEARCHER, UserRole.DATA_ADMIN, UserRole.SYS_ADMIN
                         )
                    ]
                    _raise_invalid_value(
                        _('权限"%(val)s"不可设置，可选的权限为：%(roles)s') % {
                            'val': value, 'roles': ', '.join(available_roles)
                        },
                        cell
                    )
            elif header == Header.REAL_NAME:
                _check_none(value, header, cell)
                cur_user.real_name = value
            elif header == Header.EMAIL:
                _check_none(value, header, cell)
                # regex
                if not EMAIL_REGEX.match(value):
                    _raise_invalid_value(_('邮箱格式不正确'), cell)
                if value in email_row_map:
                    _raise_invalid_value(_('邮箱与第%s行重复') % email_row_map[value], cell)
                email_row_map[value] = cell.row
                if User.objects.filter(email=value).exists():
                    _raise_invalid_value(_('系统中已存在邮箱"%s"') % value, cell)
                cur_user.email = value
            elif header == Header.INSTITUTION:
                _check_none(value, header, cell)
                cur_user.institution = value
            elif header == Header.TELEPHONE:
                cur_user.tel = value

    with transaction.atomic():
        User.objects.bulk_create(users)
    refreshed_users = User.objects.filter(username__in=[user.username for user in users])
    return refreshed_users
